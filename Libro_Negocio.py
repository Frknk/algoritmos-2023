import openpyxl
from Libro import Libro
class Libro_Negocio:
    def __init__(self):
        """
        Inicializa la clase Libro_Negocio.

        Atributos:
            listado_libros (list): Lista para almacenar objetos Libro.
            registro_libros (str): Nombre del archivo XLSX para almacenar los libros.
        """
        self.listado_libros = []
        self.registro_libros = "LIBROS_EXCEL.xlsx"

    def agregar_libro(self, codigo_libro, titulo, año, tomo, autor):
        """
        Agrega un nuevo libro a la lista y guarda en el archivo XLSX.

        Args:
            codigo_libro (str): Código del libro.
            titulo (str): Título del libro.
            año (int): Año de publicación del libro.
            tomo (str): Tomo o volumen del libro.
            autor (str): Autor del libro.
        """
        libro = Libro(codigo_libro, titulo, año, tomo, autor)
        self.listado_libros.append(libro)
        self.guardar_en_xlsx()

    def editar_libro(self, codigo_libro, nuevo_codigo, nuevo_titulo, nuevo_año, nuevo_tomo, nuevo_autor):
        """
        Edita la información de un libro existente en la lista y guarda en el archivo XLSX.

        Args:
            codigo_libro (str): Código del libro a editar.
            nuevo_codigo (str): Nuevo código del libro.
            nuevo_titulo (str): Nuevo título del libro.
            nuevo_año (int): Nuevo año de publicación del libro.
            nuevo_tomo (str): Nuevo tomo o volumen del libro.
            nuevo_autor (str): Nuevo autor del libro.

        Returns:
            bool: True si se encontró y editó el libro, False si no se encontró.
        """
        for libro in self.listado_libros:
            if libro.codigo_libro == codigo_libro:
                libro.codigo_libro = nuevo_codigo
                libro.titulo = nuevo_titulo
                libro.año = nuevo_año
                libro.tomo = nuevo_tomo
                libro.autor = nuevo_autor
                self.guardar_en_xlsx()
                return True  # Se encontró y editó el libro
        return False  # No se encontró el libro

    def eliminar_libro(self, codigo_libro):
        """
        Elimina un libro de la lista y guarda en el archivo XLSX.

        Args:
            codigo_libro (str): Código del libro a eliminar.

        Returns:
            bool: True si se encontró y eliminó el libro, False si no se encontró.
        """
        for libro in self.listado_libros:
            if libro.codigo_libro == codigo_libro:
                self.listado_libros.remove(libro)
                self.guardar_en_xlsx()
                return True  # Se encontró y eliminó el libro
        return False  # No se encontró el libro

    def cargar_desde_xlsx(self):
        """
        Carga la lista de libros desde un archivo XLSX.

        Raises:
            FileNotFoundError: Si el archivo XLSX no existe.
        """
        try:
            workbook = openpyxl.load_workbook(self.registro_libros)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                codigo_libro, titulo, año, tomo, autor = row[:5]
                libro = Libro(codigo_libro, titulo, año, tomo, autor)
                self.listado_libros.append(libro)

            workbook.close()
        except FileNotFoundError:
            print(f"El archivo {self.registro_libros} no existe.")

    def guardar_en_xlsx(self):
        """
        Guarda la lista de libros en un archivo XLSX.

        Raises:
            PermissionError: Si no se puede guardar el archivo debido a permisos.
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["CÓDIGO", "TÍTULO", "AÑO", "TOMO", "AUTOR"])

        for libro in self.listado_libros:
            sheet.append([libro.codigo_libro, libro.titulo, libro.año, libro.tomo, libro.autor])

        try:
            workbook.save(self.registro_libros)
            print("Los libros han sido guardados en el archivo XLSX.")
        except PermissionError:
            print(f"No se puede guardar el archivo {self.registro_libros}. Asegúrese de que no esté abierto en otro programa.")