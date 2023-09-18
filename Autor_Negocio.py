import openpyxl
from Autor import Autor  

class Autor_Negocio:
    def __init__(self):
        """
        Inicializa la clase Autor_Negocio.

        Atributos:
            listado_autores (list): Lista para almacenar objetos Autor.
            registro_autores (str): Nombre del archivo Excel para el registro de autores.
        """
        self.listado_autores = []
        self.registro_autores = "autores.xlsx"

    def agregar_autor(self, cod_autor, nombre, apellido_paterno, apellido_materno, fecha_nacimiento, pais, editorial):
        """
        Agrega un nuevo autor a la lista y guarda en el archivo XLSX.

        Args:
            cod_autor (str): Código del autor.
            nombre (str): Nombre del autor.
            apellido_paterno (str): Apellido paterno del autor.
            apellido_materno (str): Apellido materno del autor.
            fecha_nacimiento (str): Fecha de nacimiento del autor (en formato texto).
            pais (str): País de origen del autor.
            editorial (str): Editorial asociada al autor.
        """
        autor = Autor(cod_autor, nombre, apellido_paterno, apellido_materno, fecha_nacimiento, pais, editorial)
        self.listado_autores.append(autor)
        self.guardar_en_xlsx()

    def editar_autor(self, cod_autor, nuevo_codigo, nuevo_nombre, nuevo_apellido_paterno, nuevo_apellido_materno,
                     nueva_fecha_nacimiento, nuevo_pais, nueva_editorial):
        """
        Edita la información de un autor existente en la lista y guarda en el archivo XLSX.

        Args:
            cod_autor (str): Código del autor a editar.
            nuevo_codigo (str): Nuevo código del autor.
            nuevo_nombre (str): Nuevo nombre del autor.
            nuevo_apellido_paterno (str): Nuevo apellido paterno del autor.
            nuevo_apellido_materno (str): Nuevo apellido materno del autor.
            nueva_fecha_nacimiento (str): Nueva fecha de nacimiento del autor (en formato texto).
            nuevo_pais (str): Nuevo país de origen del autor.
            nueva_editorial (str): Nueva editorial asociada al autor.

        Returns:
            bool: True si se encontró y editó el autor, False si no se encontró.
        """
        for autor in self.listado_autores:
            if autor.cod_autor == cod_autor:
                autor.cod_autor = nuevo_codigo
                autor.nombre = nuevo_nombre
                autor.apellido_paterno = nuevo_apellido_paterno
                autor.apellido_materno = nuevo_apellido_materno
                autor.fecha_nacimiento = nueva_fecha_nacimiento
                autor.pais = nuevo_pais
                autor.editorial = nueva_editorial
                self.guardar_en_xlsx()
                return True  # Se encontró y editó el autor
        return False  # No se encontró el autor

    def eliminar_autor(self, cod_autor):
        """
        Elimina un autor de la lista y guarda en el archivo XLSX.

        Args:
            cod_autor (str): Código del autor a eliminar.

        Returns:
            bool: True si se encontró y eliminó el autor, False si no se encontró.
        """
        for autor in self.listado_autores:
            if autor.cod_autor == cod_autor:
                self.listado_autores.remove(autor)
                self.guardar_en_xlsx()
                return True  # Se encontró y eliminó el autor
        return False  # No se encontró el autor

    def cargar_desde_xlsx(self):
        """
        Carga la lista de autores desde un archivo XLSX.

        Raises:
            FileNotFoundError: Si el archivo XLSX no existe.
        """
        try:
            workbook = openpyxl.load_workbook(self.registro_autores)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                cod_autor, nombre, apellido_paterno, apellido_materno, fecha_nacimiento, pais, editorial = row[:7]
                autor = Autor(cod_autor, nombre, apellido_paterno, apellido_materno, fecha_nacimiento, pais, editorial)
                self.listado_autores.append(autor)

            workbook.close()
        except FileNotFoundError:
            print(f"El archivo {self.registro_autores} no existe.")

    def guardar_en_xlsx(self):
        """
        Guarda la lista de autores en un archivo XLSX.

        Raises:
            PermissionError: Si no se puede guardar el archivo debido a permisos.
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["CÓDIGO", "NOMBRE", "APELLIDO PATERNO", "APELLIDO MATERNO", "FECHA DE NACIMIENTO", "PAÍS", "EDITORIAL"])

        for autor in self.listado_autores:
            sheet.append([autor.cod_autor, autor.nombre, autor.apellido_paterno, autor.apellido_materno,
                          autor.fecha_nacimiento, autor.pais, autor.editorial])

        try:
            workbook.save(self.registro_autores)
            print("Los autores han sido guardados en el archivo XLSX.")
        except PermissionError:
            print(f"No se puede guardar el archivo {self.registro_autores}. Asegúrese de que no esté abierto en otro programa.")
